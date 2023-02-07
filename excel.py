import logging

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font


class ExcelUtils:
    def __init__(self):
        logging.info('EXCEL UTIL LOADED')

        self.path_excel_config = None
        self.path_excel = None
        self.workbook = None
        self.worksheet = None
        self.sheet = None

    def set_path_excel(self, path_excel):
        self.path_excel = path_excel

    def open_workbook(self, sheet_name=0):
        self.workbook = openpyxl.load_workbook(self.path_excel)
        self.worksheet = self.workbook[sheet_name]

    def create_workbook(self):
        self.workbook = openpyxl.Workbook()

    def create_sheet(self, name, index):
        logging.info("CREATING SHEET %s(%s)", str(name), str(index))
        self.sheet = self.workbook.create_sheet(name, index)

    def write_rows(self, data):
        """
        :param data: Must be list of lists
        :return:
        """
        for row in data:
            self.sheet.append(row)

    def save_workbook(self, file_name):
        logging.info("SAVING FILE %s", file_name)
        self.workbook.save(file_name)

    def get_rows(self):
        return self.worksheet.values

    def load_config_options(self):
        key = None
        value = None

        logging.info("READING OPTIONS FROM CONFIG FILE")
        for row in self.worksheet.values:
            for value in row:
                if key is None:
                    key = value
                elif value and key:
                    if "pass" not in key.lower():
                        logging.info("%s: %s", key, value)
                    setattr(self, str.lower(key), value)
                    key = None

    def remove_white_sheet(self):
        self.workbook.remove(self.workbook["Sheet"])

    def auto_width_columns(self):
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except AttributeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.sheet.column_dimensions[column].width = adjusted_width

    def set_format_to_excel(self):
        self.sheet.merge_cells("A9:D9")
        self.set_cells_format()

    def set_cells_format(self):
        white = PatternFill(start_color='FFFFFFFF',
                            end_color='FFFFFFFF',
                            fill_type='solid')
        black = PatternFill(start_color='00000000',
                            end_color='00000000',
                            fill_type='solid')
        gray = PatternFill(start_color='44444444',
                           end_color='44444444',
                           fill_type='solid')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        font = Font(name='Calibri', color="FFFFFFFF", bold=True)

        for row in self.sheet.rows:
            for cell in row:
                cell.fill = white
                if cell.row == 9:
                    cell.fill = black
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = font
                elif cell.row == 11:
                    cell.fill = gray
                    cell.border = thin_border
                    cell.font = font
                elif cell.row > 11:
                    cell.border = thin_border

    def set_image_to_excel(self):
        img = openpyxl.drawing.image.Image('image.png')
        img.anchor = "A1"
        self.sheet.add_image(img)

    def set_header_options(self, title: list, headers: list, space_rows: int = 0):
        for _ in range(space_rows):
            self.sheet.append([])
        row = [title, [], headers]
        for cell in row:
            self.sheet.append(cell)
