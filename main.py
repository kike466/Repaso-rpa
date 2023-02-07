import re
import herramienta
import requests
import openpyxl
from openpyxl import Workbook

def carrito(license):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8,gl;q=0.7,pt;q=0.6',
        'Connection': 'keep-alive',
        'Referer': 'https://www.logismarket.es/',
        'Sec-Fetch-Dest': 'iframe',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'cross-site',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Mobile Safari/537.36',
        'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
        'sec-ch-ua-mobile': '?1',
        'sec-ch-ua-platform': '"Android"',
    }
    params = {
        'license_id': license,
        'group': '11',
        'embedded': '1',
        'widget_version': '3',
        'unique_groups': '0',
    }
    response = requests.get('https://secure.livechatinc.com/customer/action/open_chat', params=params, headers=headers)

    headers = {
        'authority': 'www.logismarket.es',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-language': 'es-ES,es;q=0.9,en;q=0.8,gl;q=0.7,pt;q=0.6',
        'cookie': 'wp_customerGroup=NOT%20LOGGED%20IN; CookieConsent={stamp:%27DV3/Zg+7rl73CuYSwSJp/sP42UdfIJE6NpzZrpHcFgG8gOlKWkj8Mw==%27%2Cnecessary:true%2Cpreferences:true%2Cstatistics:true%2Cmarketing:true%2Cmethod:%27explicit%27%2Cver:1%2Cutc:1675327632562%2Cregion:%27es%27}; PHPSESSID=pgikltdg43itl7lpa9egifn8mk; form_key=C1ktFK4MqTvO1DSM; _gid=GA1.2.910510286.1675685913; _gaclientid=890189903.1675327267; _gasessionid=20230206|01006052; mage-cache-storage=%7B%7D; mage-cache-storage-section-invalidation=%7B%7D; mage-cache-sessid=true; mage-messages=; form_key=C1ktFK4MqTvO1DSM; recently_viewed_product=%7B%7D; recently_viewed_product_previous=%7B%7D; recently_compared_product=%7B%7D; recently_compared_product_previous=%7B%7D; product_data_storage=%7B%7D; _ga=GA1.2.890189903.1675327267; private_content_version=69361ee555689bc8f221b51c800121e7; X-Magento-Vary=e9c8910ceb999117606450b176c518c1fd7e44a2; section_data_ids=%7B%22customer-gtm-variables%22%3A1675687001%2C%22cart%22%3A1675687004%2C%22directory-data%22%3A1675687004%2C%22ammessages%22%3A1675687004%2C%22wp_ga4%22%3A1675687004%2C%22gtm%22%3A1675687004%2C%22ajaxpro-cart%22%3A1000%7D; _gahitid=13:57:18; _dc_gtm_UA-13304812-25=1; _dc_gtm_UA-13304812-1=1; _ga_4E5QH3DEFB=GS1.1.1675685912.4.1.1675688238.60.0.0; _ga_B5NSVWCXE7=GS1.1.1675685912.4.1.1675688238.60.0.0',
        'referer': 'https://www.logismarket.es/catalogsearch/result/?q=M0062727',
        'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
        'sec-ch-ua-mobile': '?1',
        'sec-ch-ua-platform': '"Android"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Mobile Safari/537.36',
    }

    response = requests.get('https://www.logismarket.es/checkout/cart/', headers=headers)

    print(response)

wb = openpyxl.load_workbook("Repaso.xlsx")
sheet = wb.worksheets[0]

start_row = 2
row_counter = 1
column = []
for row in sheet.iter_rows(values_only=True):
    if row_counter < start_row:
        row_counter += 1
        continue
    column.append(row[0])
wb.close()
# print(column)

nombres = []
precios = []
headers = {
    'authority': 'www.logismarket.es',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,/;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-language': 'es,es-ES;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'cache-control': 'no-cache',
    'pragma': 'no-cache',
    'sec-ch-ua': '"Not_A Brand";v="99", "Microsoft Edge";v="109", "Chromium";v="109"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.70',
}

response = requests.get('https://www.logismarket.es/', headers=headers)

# busqueda de M0062727

headers = {
    'authority': 'www.logismarket.es',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-language': 'es-ES,es;q=0.9,en;q=0.8,gl;q=0.7,pt;q=0.6',
    'cookie': 'PHPSESSID=73t7fruv03e33a5mljh1k0nk01; wp_customerGroup=NOT%20LOGGED%20IN; form_key=3Tz9TSIYrcmw1Bw7; mage-cache-storage=%7B%7D; mage-cache-storage-section-invalidation=%7B%7D; mage-cache-sessid=true; mage-messages=; recently_viewed_product=%7B%7D; recently_viewed_product_previous=%7B%7D; recently_compared_product=%7B%7D; recently_compared_product_previous=%7B%7D; product_data_storage=%7B%7D; form_key=3Tz9TSIYrcmw1Bw7; section_data_ids=%7B%22customer-gtm-variables%22%3A1675327266%7D; private_content_version=afe0eadd76f7edb36f33db3c895018f6; CookieConsent={stamp:%27DV3/Zg+7rl73CuYSwSJp/sP42UdfIJE6NpzZrpHcFgG8gOlKWkj8Mw==%27%2Cnecessary:true%2Cpreferences:true%2Cstatistics:true%2Cmarketing:true%2Cmethod:%27explicit%27%2Cver:1%2Cutc:1675327632562%2Cregion:%27es%27}; _ga=GA1.2.890189903.1675327267; _gid=GA1.2.530905109.1675327633; _gaclientid=890189903.1675327267; _gasessionid=20230202|02602108; _dc_gtm_UA-13304812-1=1; _dc_gtm_UA-13304812-25=1; _gahitid=09:47:13; _ga_4E5QH3DEFB=GS1.1.1675327266.1.1.1675327633.60.0.0; _ga_B5NSVWCXE7=GS1.1.1675327266.1.1.1675327633.60.0.0; X-Magento-Vary=e9c8910ceb999117606450b176c518c1fd7e44a2',
    'referer': 'https://www.logismarket.es/',
    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
    'sec-ch-ua-mobile': '?1',
    'sec-ch-ua-platform': '"Android"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Mobile Safari/537.36',
}

params = {
    'q': str(column[0]),
}

response = requests.get('https://www.logismarket.es/catalogsearch/result/', params=params, headers=headers)
nombre = re.search('<h2>.{0,30}</h2>', response.text)
nombre = re.split('2>', str(nombre))[1]
nombre = re.split('</', str(nombre))[0]

precio = re.search('price-excluding-tax"><span class="price">.......<', response.text)
precio = re.split('e">', str(precio))[1]
precio = re.split('\D\D', str(precio))[0]

license = re.search('license = \d{0,20}', response.text)
license = re.split('= ', str(license))[1]
license = re.split("'", str(license))[0]

carrito(license)


h1 = herramienta.Herramienta(nombre, precio)
nombres.append(nombre)
precios.append(precio)

# busqueda de M0062712

headers = {
    'authority': 'www.logismarket.es',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-language': 'es-ES,es;q=0.9,en;q=0.8,gl;q=0.7,pt;q=0.6',
    'cookie': 'PHPSESSID=73t7fruv03e33a5mljh1k0nk01; wp_customerGroup=NOT%20LOGGED%20IN; form_key=3Tz9TSIYrcmw1Bw7; mage-cache-storage=%7B%7D; mage-cache-storage-section-invalidation=%7B%7D; mage-cache-sessid=true; mage-messages=; recently_viewed_product=%7B%7D; recently_viewed_product_previous=%7B%7D; recently_compared_product=%7B%7D; recently_compared_product_previous=%7B%7D; product_data_storage=%7B%7D; form_key=3Tz9TSIYrcmw1Bw7; CookieConsent={stamp:%27DV3/Zg+7rl73CuYSwSJp/sP42UdfIJE6NpzZrpHcFgG8gOlKWkj8Mw==%27%2Cnecessary:true%2Cpreferences:true%2Cstatistics:true%2Cmarketing:true%2Cmethod:%27explicit%27%2Cver:1%2Cutc:1675327632562%2Cregion:%27es%27}; _gid=GA1.2.530905109.1675327633; _gaclientid=890189903.1675327267; _gasessionid=20230202|02602108; _dc_gtm_UA-13304812-25=1; _dc_gtm_UA-13304812-1=1; _ga=GA1.2.890189903.1675327267; _gahitid=10:08:26; _ga_B5NSVWCXE7=GS1.1.1675327266.1.1.1675328906.55.0.0; _ga_4E5QH3DEFB=GS1.1.1675327266.1.1.1675328906.55.0.0; private_content_version=ecf8542a0c3db2538ab8848d909c5f73; X-Magento-Vary=e9c8910ceb999117606450b176c518c1fd7e44a2; section_data_ids=%7B%22customer-gtm-variables%22%3A1675328902%7D',
    'referer': 'https://www.logismarket.es/',
    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
    'sec-ch-ua-mobile': '?1',
    'sec-ch-ua-platform': '"Android"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Mobile Safari/537.36',
}

params = {
    'q': str(column[1]),
}

response = requests.get('https://www.logismarket.es/catalogsearch/result/', params=params, headers=headers)
nombre = re.search('<h2>.{0,30}</h2>', response.text)
nombre = re.split('2>', str(nombre))[1]
nombre = re.split('</', str(nombre))[0]

precio = re.search('price-excluding-tax"><span class="price">.......<', response.text)
precio = re.split('e">', str(precio))[1]
precio = re.split('\D\D', str(precio))[0]

license = re.search('license = \d{0,20}', response.text)
license = re.split('= ', str(license))[1]
license = re.split("'", str(license))[0]

carrito(license)

h2 = herramienta.Herramienta(nombre, precio)

nombres.append(nombre)
precios.append(precio)

resultado = [nombre, precios]

heramientas = [h1, h2]

print(response.content)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1, value="id")
sheet.cell(row=1, column=2, value="Nombre")
sheet.cell(row=1, column=3, value="Precio")

for i, elemento in enumerate(nombres):
    sheet.cell(row=i + 2, column=1, value=column[i])
    sheet.cell(row=i + 2, column=2, value=nombres[i])
    sheet.cell(row=i + 2, column=3, value=precios[i])

wb.save("row_creation_loop.xlsx")





